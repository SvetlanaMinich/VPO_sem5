import openpyxl


COLUMNS = ['id', 'priority', 'module', 'submodule', 'steps', 'result']
FILENAME = 'testing.xlsx'

def create_excel(file_name:str):
    workbook = openpyxl.Workbook(file_name)
    # Создаем листы с названиями Task 1, Task 2, ..., Task 6
    for i in range(1, 7):
        if i == 1:
            # Первый лист уже создан по умолчанию
            sheet = workbook.active
            sheet.title = f'Task {i}'  # Название листа
        else:
            # Добавляем новые листы
            sheet = workbook.create_sheet(title=f'Task {i}')

        # Записываем заголовки в первую строку
        for col_num, column_title in enumerate(COLUMNS, 1):  # Нумерация с 1 для Excel
            sheet.cell(row=1, column=col_num, value=column_title)

    # Сохраняем Excel файл
    workbook.save(file_name)
    print(f"Excel файл '{file_name}' с колонками успешно создан!")


def fill_data(sheet, test_data=None):
    # Определяем, какая строка будет следующей для ввода новых данных
    next_row = sheet.max_row + 1

    # Заполняем строку с тестовыми данными
    for col_num, column_title in enumerate(COLUMNS, 1):
        sheet.cell(row=next_row, column=col_num, value=test_data[column_title])


def first(file_name:str):
    workbook = openpyxl.load_workbook(file_name)

    data = {
        'id' : 'task1-2',
        'priority' : '',
        'module' : 'Вывод надписей',
        'submodule' : 'Проверка вывода нужного количества восклицательных знаков в 3-ей строке',
        'steps' : '1. Запуск программы',
        'result' : '1. Вывод на экран 3х строк, каждая на новой строке:\n"Hello, world!",\n"Andhiagain!",\n"!!!!!!!!!!!!!!!!!!!!!!!!!"\nКоличество восклицательных знаков: 25'
    }

    if 'Task 1' in workbook.sheetnames:
        sheet = workbook['Task 1']
    
    fill_data(sheet, data)

    workbook.save(file_name)
    print(f"Данные успешно добавлены в Excel файл '{file_name}'!")


def second(file_name:str):
    workbook = openpyxl.load_workbook(file_name)

    data = {
        'id' : 'task2-1',
        'priority' : '',
        'module' : 'Проверка правильного выполнения задания 2',
        'submodule' : 'Проверка вывода при верном вводе данных (1ого человека)',
        'steps' : '1.Запускаем программу\n2.Вводим имя: Иван, нажимаем Enter\n3.Вводим фамилию: Иванов, нажимаем Enter\n4.Вводим возраст: 22, нажимаем Enter\n5.Нажимаем Enter для завершения ввода',
        'result' : '1.На экран выводится: "Введите Имя с заглавной буквы или нажмите <Enter> чтобы закончить: "\n2.На экран выводится: "Введите Фамилию с заглавной буквы: "\n3.На экран выводится: "Введите возраст: "\n4.На экран выводится: "Введите Имя с заглавной буквы или нажмите <Enter> чтобы закончить: "\n5.На экран выводится: "Иванов Иван 22\nМинимальный возраст: 22, Максимальный возраст: 22, Средний возраст: 22.0"'
    }

    if 'Task 2' in workbook.sheetnames:
        sheet = workbook['Task 2']
    
    fill_data(sheet, data)

    workbook.save(file_name)
    print(f"Данные успешно добавлены в Excel файл '{file_name}'!")


second(FILENAME)